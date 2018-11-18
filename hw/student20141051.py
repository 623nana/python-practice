#!/usr/bin/python3

from openpyxl import *

wb = load_workbook(filename='student.xlsx')
ws = wb.active

i = 2
grade = []

for row in ws.iter_rows():
    if row[0].value == 'id':
        continue
    midterm = int(row[2].value)
    final = int(row[3].value)
    hw = int(row[4].value)
    attendance = int(row[5].value)
    total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance
    row[6].value = total
    grade.append(float(total))

wb.save('student.xlsx')

grade.sort(reverse=True)

aPlus = int(len(grade)*0.3*0.5)-1
a = int(len(grade)*0.3)-1
b = int(len(grade)*0.7)-1
bPlus = a+int((b-a)*0.5)
c = b+int(len(grade)*0.3)
cPlus = b+int((c-b)*0.5)

while True:
    if not ws.cell(row = i, column = 1).value:
        break
    total = float(ws.cell(row = i, column = 7).value)
    if total < 40:
        ws.cell(row = i, column = 8, value="F")
    elif total >= grade[aPlus]:
        ws.cell(row = i, column = 8, value="A+")
    elif total >= grade[a]:
        ws.cell(row = i, column = 8, value="A0")
    elif total >= grade[bPlus]:
        ws.cell(row = i, column = 8, value="B+")
    elif total >= grade[b]:
        ws.cell(row = i, column = 8, value="B0")
    elif total >= grade[cPlus]:
        ws.cell(row = i, column = 8, value="C+")
    else:
        ws.cell(row = i, column = 8, value="C0")
    i = i+1

wb.save('student.xlsx')
